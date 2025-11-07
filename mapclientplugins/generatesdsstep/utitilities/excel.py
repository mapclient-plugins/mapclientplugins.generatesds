import io
import zipfile

import openpyxl
import pandas as pd
import requests

SUBJECT_ID_COLUMN = "subject id"
# Pre-defining sheet name for openpyxl.
SHEET_NAME = "Sheet1"


def get_subject_list(filepath):
    """
    Reads the Excel file and returns a clean list of subject IDs.
    """
    try:
        df = pd.read_excel(filepath, sheet_name=SHEET_NAME)

        if SUBJECT_ID_COLUMN not in df.columns:
            print(f"Error: Column '{SUBJECT_ID_COLUMN}' not found in file.")
            return []

        # Get all values from the column, drop any empty/NaN rows,
        # and convert to a standard Python list.
        subjects = df[SUBJECT_ID_COLUMN].dropna().tolist()
        return subjects

    except FileNotFoundError:
        print(f"Error: File not found at {filepath}")
        return []
    except Exception as e:
        print(f"An error occurred reading the file: {e}")
        return []


def remove_subject_by_id(subject_id_to_delete, filepath):
    """
    Finds a subject by its ID and deletes the entire row
    while preserving all Excel formatting.
    """
    wb = None
    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb[SHEET_NAME]

        row_to_delete = None

        # Find the column index for 'subject-ID'
        id_col_index = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == SUBJECT_ID_COLUMN:
                id_col_index = col
                break

        if id_col_index is None:
            print(f"Error: Could not find column '{SUBJECT_ID_COLUMN}' in sheet.")
            wb.close()
            return False

        # Find the row to delete by looking for the ID in that column
        # Start from row 2 to skip the header
        for row in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=id_col_index).value
            if cell_value == subject_id_to_delete:
                row_to_delete = row
                break

        # If we found the row, delete it and save the file
        if row_to_delete:
            ws.delete_rows(row_to_delete)
            wb.save(filepath)
            print(f"Successfully removed subject '{subject_id_to_delete}'.")
            return True
        else:
            print(f"Error: Could not find subject '{subject_id_to_delete}'.")
            return False

    except FileNotFoundError:
        print(f"Error: File not found at {filepath}")
        return False
    except KeyError:
        print(f"Error: Sheet '{SHEET_NAME}' not found in the Excel file.")
        return False
    except Exception as e:
        print(f"An error occurred removing the subject: {e}")
        return False
    finally:
        # Ensure the workbook is always closed if it was opened
        if 'wb' in locals() and wb is not None:
            wb.close()


def resolve_doi_to_url(doi_string):
    """
    Resolves a DOI to its final landing page URL.
    """
    try:
        response = requests.head(doi_string, allow_redirects=True)

        response.raise_for_status()

        final_url = response.url
        parts = final_url.split('/')
        dataset_id = parts[4]
        version_number = parts[6]
        json = {
            "data": {
                "paths": ["files/subjects.xlsx"],
                "datasetId": dataset_id,
                "version": version_number,
            }
        }

        # download the files with zipit service, single file so we actually get the content???
        api_url = "https://api.pennsieve.io/zipit/discover"
        headers = {"content-type": "application/json"}

        return api_url, headers, json

    except requests.exceptions.RequestException as e:
        print(f"Error resolving DOI: {e}")
        return None


def check_file_type(content_stream):
    """
    Checks if a byte stream is a generic zip or an xlsx file.

    Returns: "xlsx", "zip", or "unknown"
    """
    if not content_stream.startswith(b'PK\x03\x04'):
        return "unknown"

    try:
        file_bytes = io.BytesIO(content_stream)

        with zipfile.ZipFile(file_bytes, 'r') as zip_file:
            file_list = zip_file.namelist()

            if '[Content_Types].xml' in file_list or \
                    'xl/workbook.xml' in file_list:
                return "xlsx"
            else:
                return "zip"

    except zipfile.BadZipFile:
        return "unknown"
    except Exception as e:
        return "unknown but not a zip file"


def check_subject_id_is_in_dataset(subject_id_to_check: str, doi: str) -> bool:
    """
    Checks if a subject-id exists in the remote subjects.xlsx file.

    Args:
        subject_id_to_check: The subject ID string you want to find.
        doi: The URL where the subject.xlsx file is located.

    Returns:
        True if the ID is found, False otherwise.
    """

    try:
        url, headers, json = resolve_doi_to_url(doi)

        response = requests.post(url, json=json, headers=headers)
        response.raise_for_status()  # Raises an error for bad responses (404, 500, etc.)

        file_type = check_file_type(response.content)
        if file_type != "xlsx":
            return False

        file_data = io.BytesIO(response.content)

        df = pd.read_excel(file_data)

        id_column = "subject id"
        if id_column not in df.columns:
            return False

        subject_list = df[id_column].dropna().tolist()

        if subject_id_to_check in subject_list:
            return True
        else:
            return False

    except requests.exceptions.RequestException as e:
        return False
    except FileNotFoundError:
        # This can be raised by pd.read_excel if the URL was a local path
        return False
    except Exception as e:
        return False


def add_row_to_excel(data_dict, filepath, sheet_name="Sheet1"):
    """
    Appends a new row of data to an Excel file from a dictionary.

    This function preserves existing formatting by using openpyxl.
    It assumes data_dict keys EXACTLY match the header names in the
    Excel file's first row.
    """
    wb = None
    try:
        # Load the workbook and the specific sheet
        wb = openpyxl.load_workbook(filepath)
        ws = wb[sheet_name]

        header_map = {}
        for cell in ws[1]:  # Iterate over the first row (headers)
            if cell.value:
                header_map[cell.value] = cell.column

        new_row_index = ws.max_row + 1

        for header_key, value in data_dict.items():
            if header_key in header_map:
                col_index = header_map[header_key]
                # Write the value to the cell in the new row
                ws.cell(row=new_row_index, column=col_index).value = value
            else:
                print(f"Warning: Header '{header_key}' from your data "
                      "was not found in the Excel file. It will be skipped.")

        wb.save(filepath)
        return True

    except FileNotFoundError:
        return False
    except KeyError:
        return False
    except PermissionError:
        return False
    except Exception as e:
        return False
    finally:
        # Ensure the workbook is always closed if it was opened
        if 'wb' in locals() and wb is not None:
            wb.close()
